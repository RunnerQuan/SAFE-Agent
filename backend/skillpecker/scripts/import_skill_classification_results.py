from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
LIBRARY_ROOT = PROJECT_ROOT / "malicious-skill-library"
WORKBOOK_PATH = LIBRARY_ROOT / "skill_classification_table.xlsx"
SHEET_NAME = "skill_classification_table"

COL = {
    "decision_level": 0,
    "primary_group": 1,
    "risk_summary": 2,
    "risk_type_1": 3,
    "risk_type_2": 4,
    "risk_type_3": 5,
    "risk_type_4": 6,
    "platform": 7,
    "scan_level": 8,
    "business_category": 9,
    "skill_name": 10,
    "related_file_count": 11,
    "source_file": 12,
    "problem_summary": 13,
    "raw_verdict": 14,
    "cleaned_description": 25,
}

FLAG_COLS = {
    "external_server": 15,
    "api_key": 16,
    "autonomous_execution": 17,
    "browser_session_or_credentials": 18,
    "filesystem_access": 19,
    "shell_or_command_execution": 20,
    "driver_or_installer": 21,
    "credential_request": 22,
    "obfuscation_or_evasion": 23,
    "persistence_or_external_storage": 24,
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def strip_verdict_prefix(value: str) -> str:
    text = normalize_text(value)
    match = re.match(r"^[A-Z_]+:\s*[^-]+-\s*(.*)$", text)
    if match:
        return match.group(1).strip()
    return text


def is_duplicate_text(primary: str, secondary: str) -> bool:
    left = normalize_text(primary)
    right = normalize_text(secondary)
    if not left or not right:
        return False

    left_variants = {left.casefold(), strip_verdict_prefix(left).casefold()}
    right_variants = {right.casefold(), strip_verdict_prefix(right).casefold()}
    return bool(left_variants & right_variants)


def slugify(value: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "-", value.strip().casefold())
    return cleaned.strip("-")


def as_bool_flag(value: Any) -> bool:
    return normalize_text(value).upper() == "Y"


def unique_strings(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = normalize_text(value)
        key = cleaned.casefold()
        if not cleaned or key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
    return result


def get_risk_types(row: tuple[Any, ...]) -> list[str]:
    values: list[str] = []
    risk_summary = normalize_text(row[COL["risk_summary"]])
    if risk_summary:
        values.extend(part.strip() for part in risk_summary.split("|"))
    for key in ("risk_type_1", "risk_type_2", "risk_type_3", "risk_type_4"):
        value = normalize_text(row[COL[key]])
        if value:
            values.append(value)
    return unique_strings(values)


def workbook_headers(sheet) -> list[str]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return [str(value) if value is not None else f"column_{index}" for index, value in enumerate(header_row, start=1)]


def english_record(row_number: int, row: tuple[Any, ...]) -> dict[str, Any]:
    flags = {name: as_bool_flag(row[index]) for name, index in FLAG_COLS.items()}
    problem_summary = normalize_text(row[COL["problem_summary"]])
    raw_verdict = normalize_text(row[COL["raw_verdict"]])
    cleaned_description = normalize_text(row[COL["cleaned_description"]])

    if is_duplicate_text(cleaned_description, raw_verdict):
        cleaned_description = ""

    return {
        "decision_level": normalize_text(row[COL["decision_level"]]).upper(),
        "primary_group": normalize_text(row[COL["primary_group"]]),
        "risk_summary": normalize_text(row[COL["risk_summary"]]),
        "risk_types": get_risk_types(row),
        "platform": normalize_text(row[COL["platform"]]),
        "scan_level": normalize_text(row[COL["scan_level"]]),
        "business_category": normalize_text(row[COL["business_category"]]) or "Unknown",
        "skill_name": normalize_text(row[COL["skill_name"]]),
        "related_file_count": normalize_text(row[COL["related_file_count"]]),
        "source_file": normalize_text(row[COL["source_file"]]),
        "problem_summary": problem_summary,
        "raw_verdict": raw_verdict,
        "cleaned_description": cleaned_description,
        "flags": flags,
    }


def raw_record(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {headers[index]: row[index] for index in range(len(headers))}


def load_records() -> tuple[list[str], dict[str, list[dict[str, Any]]]]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    sheet = workbook[SHEET_NAME]
    headers = workbook_headers(sheet)
    records_by_dir: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        english = english_record(row_number, row)
        dir_name = f"{english['business_category']}__{english['skill_name']}"
        records_by_dir[dir_name].append(
            {
                "classification": english,
                "raw_values": raw_record(headers, row),
            }
        )

    return headers, records_by_dir


def verdict_label(records: list[dict[str, Any]]) -> str:
    levels = {item["classification"]["decision_level"] for item in records}
    if "MALICIOUS" in levels:
        return "malicious"
    if "SUSPICIOUS" in levels:
        return "unsafe"
    if "OVERREACH" in levels:
        return "mixed_risk"
    return "insufficient_evidence"


def finding_payload(record: dict[str, Any]) -> dict[str, Any]:
    item = record["classification"]
    summary = item["problem_summary"] or item["cleaned_description"] or item["raw_verdict"]
    impact = item["raw_verdict"] if not is_duplicate_text(item["raw_verdict"], summary) else ""
    return {
        "id": "",
        "cat": slugify(item["risk_types"][0]) if item["risk_types"] else "",
        "class": slugify(item["primary_group"]) if item["primary_group"] else "",
        "severity": "",
        "conf": "",
        "summary": summary,
        "arts": [],
        "spans": [],
        "impact": impact,
        "fix": "",
        "decision_level": item["decision_level"],
        "primary_group": item["primary_group"],
        "scan_level": item["scan_level"],
        "risk_types": item["risk_types"],
        "source_platform": item["platform"],
        "source_file": item["source_file"],
        "related_file_count": item["related_file_count"],
        "flags": item["flags"],
    }


def build_payload(skill_dir: Path, headers: list[str], records: list[dict[str, Any]]) -> dict[str, Any]:
    findings = [finding_payload(record) for record in records]
    claims = unique_strings(
        [
            record["classification"]["cleaned_description"]
            or record["classification"]["problem_summary"]
            or record["classification"]["raw_verdict"]
            for record in records
        ]
    )
    issue_types = unique_strings(
        [
            *[risk for record in records for risk in record["classification"]["risk_types"]],
            *[record["classification"]["primary_group"] for record in records],
        ]
    )
    high_risk_features = sorted(
        {
            key
            for record in records
            for key, value in record["classification"]["flags"].items()
            if value
        }
    )
    top_findings = [
        {
            "id": finding["id"],
            "cat": finding["cat"],
            "class": finding["class"],
            "severity": finding["severity"],
            "conf": finding["conf"],
            "summary": finding["summary"],
            "impact": finding["impact"],
            "fix": finding["fix"],
            "decision_level": finding["decision_level"],
            "primary_group": finding["primary_group"],
            "scan_level": finding["scan_level"],
            "risk_types": finding["risk_types"],
            "source_platform": finding["source_platform"],
            "source_file": finding["source_file"],
            "related_file_count": finding["related_file_count"],
            "flags": finding["flags"],
            "spans": [],
        }
        for finding in findings
    ]

    payload = {
        "preprocess": {
            "skill": {
                "root": str(skill_dir),
                "claims": claims,
                "observed_kinds": ["spreadsheet", "imported_scan_result"],
            },
            "artifacts": [],
            "rule_hits": [],
            "coverage_gaps": [],
            "stats": {
                "artifact_count": 0,
                "rule_hit_count": len(findings),
                "severity_counts": {},
                "rule_code_counts": {},
            },
        },
        "triage": {
            "skill": {
                "claims": claims,
                "obs": unique_strings([record["classification"]["scan_level"] for record in records]),
                "high_risk_features": high_risk_features,
            },
            "arts": [],
            "xconcerns": [],
            "retrieve": [],
            "coverage": {
                "total": len(records),
                "indexed": len(records),
                "gaps": [],
            },
        },
        "security": {
            "findings": findings,
            "ctx_requests": [],
            "coverage": {
                "reviewed_arts": [],
                "unresolved_high_risk_arts": [],
            },
        },
        "judge": {
            "verdict": {
                "label": verdict_label(records),
                "primary_concern": slugify(issue_types[0]) if issue_types else "",
                "issue_types": [slugify(item) for item in issue_types],
            },
            "top_findings": top_findings,
            "coverage_audit": {
                "adequate": True,
                "missed": [],
                "needs_rescan": False,
            },
            "next_action": "stop",
        },
        "excel_import": {
            "record_count": len(records),
            "headers": headers,
            "records": [
                {
                    "classification": record["classification"],
                    "raw_values": record["raw_values"],
                }
                for record in records
            ],
        },
    }
    return payload


def main() -> None:
    headers, records_by_dir = load_records()
    library_dirs = sorted(path for path in LIBRARY_ROOT.iterdir() if path.is_dir())

    missing = [path.name for path in library_dirs if path.name not in records_by_dir]
    if missing:
        raise RuntimeError(f"Missing Excel records for directories: {missing[:10]}")

    for skill_dir in library_dirs:
        payload = build_payload(skill_dir, headers, records_by_dir[skill_dir.name])
        (skill_dir / "scan-result.raw.json").write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        (skill_dir / "judge.raw.json").write_text(
            json.dumps(payload["judge"], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    print(f"Imported scan results for {len(library_dirs)} skills from {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
